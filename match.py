import random
import pandas as pd
from typing import List, Tuple, Set
from itertools import combinations
import os
import tkinter as tk
from tkinter import messagebox
import sys
import logging
import traceback
from pathlib import Path
import random
import datetime
from tkinter import filedialog

# 配置日誌系統
def setup_logging():
    """設置日誌系統，支持文件和控制台輸出"""
    try:
        # 創建日誌目錄
        log_dir = Path.home() / 'Desktop' / 'MatchMember_Logs'
        log_dir.mkdir(exist_ok=True)
        
        # 配置日誌格式
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        
        # 創建文件處理器
        log_file = log_dir / f'match_log_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter(log_format))
        
        # 創建控制台處理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(logging.Formatter(log_format))
        
        # 配置根日誌器
        logging.basicConfig(
            level=logging.DEBUG,
            handlers=[file_handler, console_handler],
            format=log_format
        )
        
        logging.info(f"日誌系統已啟動，日誌文件：{log_file}")
        return str(log_file)
    except Exception as e:
        # 如果日誌系統初始化失敗，使用標準輸出
        import sys
        print(f"日誌系統初始化失敗：{e}", file=sys.stderr)
        return None

# 重定向標準輸出和標準錯誤流
class OutputRedirector:
    def __init__(self, debug=False):
        self.debug = debug
        self.original_stdout = sys.stdout
        self.original_stderr = sys.stderr
        self.logger = logging.getLogger('OutputRedirector')
    
    def write(self, text):
        if self.debug and text.strip():
            self.original_stdout.write(text)
            # 同時記錄到日誌
            if hasattr(self, 'logger'):
                self.logger.debug(text.strip())
    
    def flush(self):
        if self.debug:
            self.original_stdout.flush()

# 初始化日誌系統
log_file_path = setup_logging()

# 將標準輸出重定向
sys.stdout = OutputRedirector(debug=True)
sys.stderr = OutputRedirector(debug=True)

class MatchingGUI:
    def __init__(self):
        self.logger = logging.getLogger('MatchingGUI')
        self.logger.info("初始化配對GUI")
        
        # 創建主視窗
        self.window = tk.Tk()
        self.window.title("人員配對系統 v2.0")
        self.window.geometry("500x400")  # 增大視窗以容納更多功能
        
        # 設置視窗圖標（如果存在）
        try:
            icon_path = Path(__file__).parent / "MyIcon.icns"
            if icon_path.exists():
                self.window.iconbitmap(str(icon_path))
        except Exception as e:
            self.logger.warning(f"無法設置視窗圖標：{e}")
        
        self.setup_ui()
        
        # 初始化變數
        self.current_excel_path = None
        
    def setup_ui(self):
        """設置用戶界面"""
        # 標題
        title_label = tk.Label(self.window, text="人員配對系統", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # 文件選擇區域
        file_frame = tk.Frame(self.window)
        file_frame.pack(pady=10, padx=20, fill=tk.X)
        
        tk.Label(file_frame, text="Excel 檔案：").pack(anchor=tk.W)
        
        path_frame = tk.Frame(file_frame)
        path_frame.pack(fill=tk.X, pady=5)
        
        self.filename_var = tk.StringVar(value="配對名單.xlsx")
        self.path_entry = tk.Entry(path_frame, textvariable=self.filename_var, width=40)
        self.path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        browse_button = tk.Button(path_frame, text="瀏覽", command=self.browse_file)
        browse_button.pack(side=tk.RIGHT, padx=(5, 0))
        
        # 提示文字
        hint_label = tk.Label(file_frame, text="提示：檔案將存放在桌面，或點擊瀏覽選擇位置", 
                             font=("Arial", 9), fg="gray")
        hint_label.pack(anchor=tk.W, pady=(0, 5))
        
        # 系統狀態區域
        status_frame = tk.Frame(self.window)
        status_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
        
        tk.Label(status_frame, text="系統狀態：").pack(anchor=tk.W)
        
        # 狀態顯示（使用文字框替代標籤）
        self.status_text = tk.Text(status_frame, height=8, width=50, wrap=tk.WORD)
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=5)
        self.status_text.config(state='disabled')  # 預設為不可編輯
        
        # 添加滾動條
        scrollbar = tk.Scrollbar(status_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.status_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.status_text.yview)
        
        # 按鈕區域
        button_frame = tk.Frame(self.window)
        button_frame.pack(pady=10)
        
        # 配對按鈕
        self.match_button = tk.Button(button_frame, text="開始配對", command=self.execute_matching,
                                     bg="#4CAF50", fg="white", font=("Arial", 12, "bold"),
                                     width=12, height=2)
        self.match_button.pack(side=tk.LEFT, padx=5)
        
        # 檢查配置按鈕
        check_button = tk.Button(button_frame, text="檢查配置", command=self.check_configuration,
                                bg="#2196F3", fg="white", font=("Arial", 12),
                                width=12, height=2)
        check_button.pack(side=tk.LEFT, padx=5)
        
        # 打開日誌按鈕
        if log_file_path:
            log_button = tk.Button(button_frame, text="查看日誌", command=self.open_log_file,
                                  bg="#FF9800", fg="white", font=("Arial", 12),
                                  width=12, height=2)
            log_button.pack(side=tk.LEFT, padx=5)
        
        # 初始狀態訊息
        self.update_status("系統已就緒，請檢查配置或開始配對")
        
    def browse_file(self):
        """瀏覽並選擇Excel文件位置"""
        try:
            initial_dir = str(Path.home() / 'Desktop')
            file_path = filedialog.asksaveasfilename(
                title="選擇Excel檔案保存位置",
                initialdir=initial_dir,
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if file_path:
                self.filename_var.set(file_path)
                self.current_excel_path = file_path
                self.logger.info(f"用戶選擇文件路徑：{file_path}")
                self.update_status(f"已選擇文件：{Path(file_path).name}")
        except Exception as e:
            self.logger.error(f"文件瀏覽失敗：{e}")
            self.update_status(f"文件瀏覽失敗：{e}", True)
            
    def check_configuration(self):
        """檢查系統配置和文件狀態"""
        try:
            self.logger.info("開始檢查系統配置")
            self.update_status("正在檢查系統配置...")
            
            config_messages = []
            
            # 檢查文件路徑
            filename = self.filename_var.get()
            if not filename.endswith('.xlsx'):
                config_messages.append("❌ 文件名必須以 .xlsx 結尾")
            else:
                config_messages.append("✅ 文件格式正確")
            
            # 確定完整路徑
            if self.current_excel_path:
                excel_path = self.current_excel_path
            else:
                desktop_path = Path.home() / 'Desktop'
                excel_path = desktop_path / filename
            
            config_messages.append(f"📁 文件路徑：{excel_path}")
            
            # 檢查文件是否存在
            if Path(excel_path).exists():
                config_messages.append("✅ Excel文件已存在")
                
                # 檢查工作表
                try:
                    excel_file = pd.ExcelFile(excel_path)
                    sheets = excel_file.sheet_names
                    
                    if '人員名單' in sheets:
                        config_messages.append("✅ 找到'人員名單'工作表")
                        
                        # 檢查人員名單內容
                        df = pd.read_excel(excel_path, sheet_name='人員名單')
                        if '姓名' in df.columns:
                            people_count = len(df['姓名'].dropna())
                            config_messages.append(f"👥 人員名單中有 {people_count} 人")
                        else:
                            config_messages.append("❌ 人員名單中缺少'姓名'欄位")
                    else:
                        config_messages.append("❌ 未找到'人員名單'工作表")
                    
                    if '參與配對人員' in sheets:
                        config_messages.append("✅ 找到'參與配對人員'工作表")
                        
                        # 檢查參與配對人員內容
                        df = pd.read_excel(excel_path, sheet_name='參與配對人員')
                        if '姓名' in df.columns:
                            participants_count = len(df['姓名'].dropna())
                            config_messages.append(f"🎯 參與配對人員有 {participants_count} 人")
                            
                            if participants_count == 0:
                                config_messages.append("⚠️ 參與配對人員名單為空，無法進行配對")
                            elif participants_count == 1:
                                config_messages.append("⚠️ 只有1人參與配對，無法進行配對")
                        else:
                            config_messages.append("❌ 參與配對人員中缺少'姓名'欄位")
                    else:
                        config_messages.append("❌ 未找到'參與配對人員'工作表")
                        
                except Exception as e:
                    config_messages.append(f"❌ 讀取Excel文件失敗：{e}")
            else:
                config_messages.append("⚠️ Excel文件不存在，將創建新文件")
            
            # 檢查日誌系統
            if log_file_path:
                config_messages.append(f"📝 日誌文件：{Path(log_file_path).name}")
            else:
                config_messages.append("⚠️ 日誌系統未啟動")
            
            # 顯示檢查結果
            result_text = "\n".join(config_messages)
            self.update_status(f"配置檢查完成：\n{result_text}")
            
            self.logger.info("系統配置檢查完成")
            
        except Exception as e:
            error_msg = f"配置檢查失敗：{e}"
            self.logger.error(error_msg)
            self.update_status(error_msg, True)
            
    def open_log_file(self):
        """打開日誌文件"""
        try:
            if log_file_path and Path(log_file_path).exists():
                if sys.platform == 'darwin':  # macOS
                    os.system(f'open "{log_file_path}"')
                elif sys.platform == 'win32':  # Windows
                    os.startfile(log_file_path)
                else:  # Linux
                    os.system(f'xdg-open "{log_file_path}"')
                self.logger.info("已打開日誌文件")
            else:
                self.update_status("日誌文件不存在", True)
        except Exception as e:
            self.logger.error(f"打開日誌文件失敗：{e}")
            self.update_status(f"打開日誌文件失敗：{e}", True)
        
    def update_status(self, message: str, is_error: bool = False):
        """更新狀態文字框的內容"""
        self.status_text.config(state='normal')  # 暫時允許編輯
        self.status_text.delete(1.0, tk.END)
        self.status_text.insert(tk.END, message)
        if is_error:
            self.status_text.config(fg="red")
        else:
            self.status_text.config(fg="green")
        self.status_text.config(state='disabled')  # 恢復為不可編輯
        
    def execute_matching(self):
        """執行配對並儲存結果"""
        try:
            self.logger.info("開始執行配對")
            self.update_status("正在準備配對...")
            
            # 禁用配對按鈕防止重複點擊
            self.match_button.config(state='disabled')
            self.window.update()
            
            filename = self.filename_var.get()
            if not filename.endswith('.xlsx'):
                error_msg = "檔案名稱必須以 .xlsx 結尾"
                self.logger.error(error_msg)
                self.update_status(f"失敗：{error_msg}", True)
                return
            
            # 確定完整路徑
            if self.current_excel_path:
                excel_path = self.current_excel_path
            else:
                desktop_path = Path.home() / 'Desktop'
                excel_path = desktop_path / filename
            
            self.logger.info(f"使用Excel文件路徑：{excel_path}")
            self.update_status(f"正在讀取文件：{Path(excel_path).name}...")
            
            # 建立配對名單實例
            matcher = MatchingSystem(str(excel_path))
            
            self.update_status("正在執行配對算法...")
            
            # 執行配對
            matches, repeated_pairs = matcher.match_people()
            
            self.logger.info(f"配對完成 - 總配對數: {len(matches)}, 重複配對數: {len(repeated_pairs)}")
            
            self.update_status("正在保存配對結果...")
            
            # 儲存結果
            matcher.save_matching_result(matches, repeated_pairs)
            
            # 準備結果訊息
            result_messages = [
                f"✅ 配對完成！",
                f"📊 總配對組數：{len(matches)}",
                f"👥 參與人數：{sum(len(match) for match in matches)}"
            ]
            
            if repeated_pairs:
                result_messages.append(f"⚠️ 重複配對：{len(repeated_pairs)} 組")
                result_messages.append("請檢查Excel文件中的黃色標記")
            else:
                result_messages.append("🎉 無重複配對！")
            
            result_messages.append(f"💾 結果已保存至：{Path(excel_path).name}")
            
            # 更新狀態
            status_text = "\n".join(result_messages)
            self.update_status(status_text)
            
            # 準備詳細結果顯示
            detail_text = "配對結果詳情：\n\n"
            for i, match in enumerate(matches, 1):
                detail_text += f"{i}. {' ↔ '.join(match)}\n"
            
            if repeated_pairs:
                detail_text += "\n重複配對警告：\n"
                for i, pair in enumerate(repeated_pairs, 1):
                    detail_text += f"{i}. {' ↔ '.join(pair)}\n"
                detail_text += "\n這些配對在歷史記錄中已存在，已在Excel中標記為黃色。"
            
            # 顯示結果對話框
            messagebox.showinfo("配對完成", detail_text)
            
            self.logger.info("配對流程完成")
            
        except FileNotFoundError as e:
            error_msg = f"找不到Excel文件：{e}"
            self.logger.error(error_msg)
            self.update_status(error_msg, True)
            messagebox.showerror("文件錯誤", "找不到指定的Excel文件，請檢查文件路徑是否正確。")
            
        except pd.errors.EmptyDataError as e:
            error_msg = f"Excel文件為空或格式錯誤：{e}"
            self.logger.error(error_msg)
            self.update_status(error_msg, True)
            messagebox.showerror("數據錯誤", "Excel文件為空或格式不正確，請檢查文件內容。")
            
        except PermissionError as e:
            error_msg = f"文件權限錯誤：{e}"
            self.logger.error(error_msg)
            self.update_status(error_msg, True)
            messagebox.showerror("權限錯誤", "無法訪問Excel文件，請檢查文件是否被其他程序占用或權限設置。")
            
        except Exception as e:
            error_msg = f"配對失敗：{str(e)}"
            self.logger.error(f"{error_msg}\n{traceback.format_exc()}")
            self.update_status(error_msg, True)
            
            # 顯示詳細錯誤信息
            error_detail = f"發生未預期的錯誤：\n{str(e)}\n\n請檢查：\n1. Excel文件格式是否正確\n2. 參與配對人員名單是否有效\n3. 查看日誌文件獲取更多信息"
            messagebox.showerror("系統錯誤", error_detail)
            
        finally:
            # 重新啟用配對按鈕
            self.match_button.config(state='normal')
    
    def run(self):
        self.window.mainloop()

class MatchingSystem:
    def __init__(self, excel_filename: str):
        self.logger = logging.getLogger(__name__)
        
        # 處理文件路徑
        if os.path.isabs(excel_filename):
            self.excel_path = excel_filename
        else:
            # 獲取桌面路徑
            self.desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            # 完整的 Excel 檔案路徑
            self.excel_path = os.path.join(self.desktop_path, excel_filename)
        
        self.logger.info(f"初始化配對系統，Excel路徑：{self.excel_path}")
        
        try:
            # 檢查文件是否存在
            if os.path.exists(self.excel_path):
                file_size = os.path.getsize(self.excel_path)
                self.logger.info(f"Excel文件存在，大小：{file_size} bytes")
                # 嘗試讀取現有的 Excel 檔案
                self.excel = pd.ExcelFile(self.excel_path)
            else:
                self.logger.warning(f"Excel文件不存在，將創建新文件：{self.excel_path}")
                raise FileNotFoundError("文件不存在")
                
        except FileNotFoundError:
            # 如果檔案不存在，創建新的 Excel 檔案
            self.logger.info("創建新的Excel文件")
            people_df = pd.DataFrame(columns=['姓名'])
            participants_df = pd.DataFrame(columns=['姓名'])
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                people_df.to_excel(writer, sheet_name='人員名單', index=False)
                participants_df.to_excel(writer, sheet_name='參與配對人員', index=False)
            
            self.logger.info("新Excel文件創建完成")
            self.excel = pd.ExcelFile(self.excel_path)
            
        except Exception as e:
            error_msg = f"初始化Excel文件時發生錯誤：{e}"
            self.logger.error(error_msg)
            raise Exception(error_msg)
    
    def get_all_people(self) -> List[str]:
        """獲取所有待配對人員名單"""
        try:
            self.logger.info("正在讀取人員名單...")
            
            if not os.path.exists(self.excel_path):
                raise FileNotFoundError(f"Excel文件不存在：{self.excel_path}")
            
            df = pd.read_excel(self.excel_path, sheet_name='人員名單')
            
            if '姓名' not in df.columns:
                raise ValueError("人員名單工作表中找不到'姓名'欄位")
            
            people = [name for name in df['姓名'].dropna().tolist() if name and str(name).strip()]
            self.logger.info(f"成功讀取 {len(people)} 位人員")
            return people
            
        except FileNotFoundError:
            raise
        except ValueError:
            raise
        except Exception as e:
            error_msg = f"讀取人員名單時發生錯誤: {e}"
            self.logger.error(error_msg)
            raise Exception(error_msg)
        
    def get_matching_history(self) -> Set[Tuple[str, ...]]:
        """從人員名單獲取歷史配對記錄"""
        history_set = set()
        
        try:
            self.logger.info("正在讀取歷史配對記錄...")
            
            if not os.path.exists(self.excel_path):
                self.logger.warning("Excel文件不存在，返回空的歷史記錄")
                return history_set
            
            # 讀取人員名單
            df = pd.read_excel(self.excel_path, sheet_name='人員名單')
            
            # 確保有「姓名」欄位
            if '姓名' not in df.columns:
                self.logger.warning("人員名單工作表中找不到'姓名'欄位")
                return history_set
            
            self.logger.info(f"檢查 Excel 檔案中的所有欄位: {df.columns.tolist()}")
            
            # 獲取所有配對者欄位（除了「姓名」以外的所有欄位）
            partner_columns = [col for col in df.columns if col != '姓名' and '配對者' in col]
            
            # 如果沒有配對者欄位，返回空集合
            if not partner_columns:
                self.logger.warning("Excel 檔案中未找到任何配對者欄位")
                return history_set
            
            # 打印檢查欄位，用於偵錯
            self.logger.info(f"找到 {len(partner_columns)} 個配對者欄位: {partner_columns}")
            
            # 遍歷每一行（每個人）
            for idx, row in df.iterrows():
                person = row['姓名']
                if not isinstance(person, str) or not person.strip():
                    continue
                    
                # 移除人名前的 @ 符號進行比較
                person_clean = person[1:].strip() if person.startswith('@') else person.strip()
                self.logger.debug(f"檢查人員: {person} -> {person_clean}")
                    
                # 遍歷該人的所有配對者
                for col in partner_columns:
                    partner = row[col]
                    if not pd.isna(partner) and isinstance(partner, str) and partner.strip():
                        # 移除配對者名字前的 @ 符號進行比較
                        partner_clean = partner[1:].strip() if partner.startswith('@') else partner.strip()
                        self.logger.debug(f"檢查配對者: {partner} -> {partner_clean}")
                        
                        if partner_clean and person_clean and partner_clean != person_clean:
                            # 將配對加入歷史記錄（使用不帶 @ 的名字）
                            pair = tuple(sorted([person_clean, partner_clean]))
                            history_set.add(pair)
                            self.logger.debug(f"添加歷史配對: {pair}")
            
            self.logger.info(f"成功讀取 {len(history_set)} 組歷史配對記錄")
            if self.logger.isEnabledFor(logging.DEBUG):
                self.logger.debug(f"完整的歷史配對清單: {history_set}")
            return history_set
            
        except Exception as e:
            error_msg = f"讀取配對歷史時出錯: {str(e)}"
            self.logger.error(f"{error_msg}\n{traceback.format_exc()}")
            return history_set
    
    def save_matching_result(self, matches: List[Tuple[str, ...]], repeated_pairs: List[Tuple[str, ...]] = None):
        """保存本次配對結果，並標記重複配對"""
        try:
            self.logger.info("=== 開始保存配對結果 ===")
            self.logger.info(f"配對組數: {len(matches)}")
            self.logger.info(f"重複配對組數: {len(repeated_pairs) if repeated_pairs else 0}")
            
            if not matches:
                self.logger.warning("沒有配對結果需要保存")
                return
            self.logger.debug(f"repeated_pairs 參數: {repeated_pairs}")
            self.logger.debug(f"matches 詳細內容: {matches}")
            
            import pandas as pd  # 確保 pd 在整個函數中可用
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            
            if repeated_pairs is None:
                repeated_pairs = []
            
            # 設定重複配對的樣式
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_font = Font(color="FF0000", bold=True)
            
            # 讀取現有的 Excel 檔案
            workbook = openpyxl.load_workbook(self.excel_path)
            
            # 創建配對結果字典，方便查詢每個人的配對者
            match_dict = {}
            for match in matches:
                if len(match) == 2:
                    # 確保每個人的鍵存在
                    if match[0] not in match_dict:
                        match_dict[match[0]] = []
                    if match[1] not in match_dict:
                        match_dict[match[1]] = []
                    
                    # 添加配對關係（確保只有一個 @ 前綴）
                    partner1 = match[1]
                    if not isinstance(partner1, str):
                        partner1_with_at = f"@{str(partner1)}"
                    elif partner1.startswith('@'):
                        partner1_with_at = partner1  # 已有 @ 前綴，保持不變
                    else:
                        partner1_with_at = f"@{partner1}"
                    
                    partner2 = match[0]
                    if not isinstance(partner2, str):
                        partner2_with_at = f"@{str(partner2)}"
                    elif partner2.startswith('@'):
                        partner2_with_at = partner2  # 已有 @ 前綴，保持不變
                    else:
                        partner2_with_at = f"@{partner2}"
                    
                    match_dict[match[0]].append(partner1_with_at)
                    match_dict[match[1]].append(partner2_with_at)
                elif len(match) == 3:
                    # 確保每個人的鍵存在
                    if match[0] not in match_dict:
                        match_dict[match[0]] = []
                    if match[1] not in match_dict:
                        match_dict[match[1]] = []
                    if match[2] not in match_dict:
                        match_dict[match[2]] = []
                    
                    # 添加配對關係（確保只有一個 @ 前綴）
                    for i in range(3):
                        for j in range(3):
                            if i != j:  # 避免自己配對自己
                                person = match[i]
                                partner = match[j]
                                
                                # 確保 partner 只有一個 @ 前綴
                                if not isinstance(partner, str):
                                    partner_with_at = f"@{str(partner)}"
                                elif partner.startswith('@'):
                                    partner_with_at = partner  # 已有 @ 前綴，保持不變
                                else:
                                    partner_with_at = f"@{partner}"
                                
                                match_dict[person].append(partner_with_at)
            
            # 輸出配對結果供檢查
            self.logger.debug(f"配對字典: {match_dict}")
            
            # 在填充 Excel 前記錄 match_dict
            self.logger.info("配對字典詳情：")
            for person, partners in match_dict.items():
                self.logger.info(f"{person}: {partners}")
            
            # 創建人員名單 DataFrame
            # 收集所有參與配對的人員（包括配對者和被配對者）
            all_people = set()
            for person, partners in match_dict.items():
                if isinstance(person, str):
                    person_clean = person[1:].strip() if person.startswith('@') else person.strip()
                    all_people.add(person_clean)
                
                # 添加所有配對者
                for partner in partners:
                    if isinstance(partner, str):
                        partner_clean = partner[1:].strip() if partner.startswith('@') else partner.strip()
                        all_people.add(partner_clean)
            
            # 使用 all_people 替代原來的方法
            people_list = sorted(list(all_people))
            self.logger.info(f"參與配對人員列表 ({len(people_list)} 人): {people_list}")
            
            import datetime
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            
            # 找出最大配對者數量
            max_partners = max([len(partners) for partners in match_dict.values()], default=1)
            self.logger.info(f"最大配對者數量: {max_partners}")
            
            # 更新人員名單工作表
            try:
                # 讀取現有的人員名單，保留歷史配對資料
                existing_people_df = pd.read_excel(self.excel_path, sheet_name='人員名單')
                self.logger.info(f"現有人員名單欄位: {existing_people_df.columns.tolist()}")
                self.logger.info(f"現有人員數量: {len(existing_people_df)}")
                
                # 獲取當前日期作為新欄位名稱
                import datetime
                today = datetime.datetime.now().strftime("%Y-%m-%d")
                
                # 創建新的配對者欄位名稱
                new_columns = []
                for i in range(max_partners):
                    if max_partners > 1:
                        new_columns.append(f"配對者 {today} {i+1}")
                    else:
                        new_columns.append(f"配對者 {today}")
                
                # 創建包含新配對的 DataFrame
                new_data = {'姓名': [f"@{person}" for person in people_list]}
                
                # 添加新的配對結果
                for col_name in new_columns:
                    new_data[col_name] = [''] * len(people_list)  # 初始化為空字串
                
                for i, person in enumerate(people_list):
                    person_with_at = f"@{person}"
                    for j, col_name in enumerate(new_columns):
                        if j < len(match_dict.get(person, [])):
                            new_data[col_name][i] = match_dict[person][j]
                        elif j < len(match_dict.get(person_with_at, [])):
                            new_data[col_name][i] = match_dict[person_with_at][j]
                
                new_df = pd.DataFrame(new_data)
                
                # 合併現有資料和新資料
                # 先確認哪些人已經存在，哪些人是新的
                existing_names = existing_people_df['姓名'].tolist()
                
                # 重新組織 DataFrame，將新的配對欄位插入在「姓名」欄位之後
                new_columns_order = ['姓名']
                new_columns_order.extend(new_columns)  # 新配對欄位
                
                # 添加其他原有欄位
                for col in existing_people_df.columns:
                    if col != '姓名' and col not in new_columns:
                        new_columns_order.append(col)
                
                # 創建新的 DataFrame
                merged_df = pd.DataFrame(columns=new_columns_order)
                
                # 初始化為空值
                for col in new_columns_order:
                    merged_df[col] = ''
                
                # 複製現有數據
                for i, row in existing_people_df.iterrows():
                    new_row = {}
                    for col in existing_people_df.columns:
                        if col in new_columns_order:
                            new_row[col] = row[col]
                    
                    # 添加到新 DataFrame
                    merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)
                
                # 更新現有人員的新配對資料
                for i, row in new_df.iterrows():
                    name = row['姓名']
                    # 檢查是否為現有人員
                    if name in existing_names:
                        # 更新現有人員的新配對
                        idx = existing_names.index(name)
                        for col in new_columns:
                            merged_df.at[idx, col] = row[col]
                    else:
                        # 添加新人員
                        new_row = pd.Series(index=new_columns_order)
                        new_row['姓名'] = name
                        for col in new_columns:
                            new_row[col] = row[col]
                        merged_df = pd.concat([merged_df, pd.DataFrame([new_row])], ignore_index=True)
                
                # 記錄 merged_df 的最終結果
                self.logger.debug("\n合併後的 DataFrame (merged_df)：")
                self.logger.debug(f"\n{merged_df}\n")
                
                # 在寫入 Excel 前，保留原始參與配對人員
                try:
                    # 先嘗試讀取現有的參與配對人員
                    existing_participants_df = pd.read_excel(self.excel_path, sheet_name='參與配對人員')
                except:
                    # 如果讀取失敗，則使用空的 DataFrame
                    existing_participants_df = pd.DataFrame(columns=['姓名'])
                
                # 直接使用 openpyxl 保存資料，避免 Pandas 修改欄位名稱
                from copy import copy
                
                # 先讀取原始的工作簿，保留所有原始格式和內容
                workbook = openpyxl.load_workbook(self.excel_path)
                
                # 如果人員名單工作表已存在，則獲取它
                if '人員名單' in workbook.sheetnames:
                    people_sheet = workbook['人員名單']
                    
                    # 找到姓名列的索引
                    name_col_idx = None
                    for col_idx, cell in enumerate(people_sheet[1], 1):
                        if cell.value == '姓名':
                            name_col_idx = col_idx
                            break
                    
                    if name_col_idx is None:
                        name_col_idx = 1
                    
                    # 獲取所有現有欄位
                    existing_cols = []
                    for col_idx, cell in enumerate(people_sheet[1], 1):
                        if cell.value:
                            existing_cols.append((col_idx, cell.value))
                    
                    # 移動所有在姓名欄之後的列，為新配對欄位騰出空間
                    # 從最右邊的列開始往右移動
                    for i in range(len(existing_cols) - 1, 0, -1):
                        if existing_cols[i][0] > name_col_idx:
                            target_col_idx = existing_cols[i][0] + len(new_columns)
                            source_col_idx = existing_cols[i][0]
                            
                            self.logger.info(f"移動欄位: 從第{source_col_idx}列移動到第{target_col_idx}列 (欄位名稱: {existing_cols[i][1]})")
                            
                            # 移動整列數據
                            for row_idx in range(1, people_sheet.max_row + 1):
                                source_cell = people_sheet.cell(row=row_idx, column=source_col_idx)
                                target_cell = people_sheet.cell(row=row_idx, column=target_col_idx)
                                
                                # 複製單元格值和樣式
                                target_cell.value = source_cell.value
                                if source_cell.has_style:
                                    target_cell.font = copy(source_cell.font)
                                    target_cell.border = copy(source_cell.border)
                                    target_cell.fill = copy(source_cell.fill)
                                    target_cell.number_format = copy(source_cell.number_format)
                                    target_cell.protection = copy(source_cell.protection)
                                    target_cell.alignment = copy(source_cell.alignment)
                                
                                # 清空原來的位置
                                source_cell.value = None
                                # 重置樣式
                                source_cell.font = openpyxl.styles.Font()
                                source_cell.border = openpyxl.styles.Border()
                                source_cell.fill = openpyxl.styles.PatternFill()
                                source_cell.number_format = 'General'
                                source_cell.alignment = openpyxl.styles.Alignment()
                    
                    # 在姓名列右側插入新的配對欄位
                    for i, col_name in enumerate(new_columns):
                        col_idx = name_col_idx + 1 + i
                        people_sheet.cell(row=1, column=col_idx).value = col_name
                    
                    # 從 match_dict 中填入配對結果
                    name_to_row_idx = {}
                    for row_idx in range(2, people_sheet.max_row + 1):
                        name = people_sheet.cell(row=row_idx, column=name_col_idx).value
                        if name:
                            # 儲存名稱和行索引的映射，便於填入配對結果
                            name_clean = name[1:].strip() if isinstance(name, str) and name.startswith('@') else name
                            name_to_row_idx[name_clean] = row_idx
                            name_to_row_idx[f"@{name_clean}"] = row_idx
                    
                    # 在創建 name_to_row_idx 映射後記錄它
                    self.logger.debug("名稱到行索引映射 (name_to_row_idx)：")
                    self.logger.debug(f"{name_to_row_idx}")
                    
                    # 填入配對結果
                    self.logger.info(f"=== 第一階段：填入已存在人員的配對結果 ===")
                    self.logger.info(f"開始填入配對結果，match_dict 有 {len(match_dict)} 個項目")
                    for person, partners in match_dict.items():
                        self.logger.debug(f"\n處理人員: {person}, 配對者: {partners}")
                        # 忽略可能的數字索引或其他非人名鍵
                        if not isinstance(person, str):
                            self.logger.debug(f"  跳過非字串鍵: {person}")
                            continue
                            
                        person_clean = person[1:].strip() if person.startswith('@') else person.strip()
                        self.logger.debug(f"  清理後的人員名稱: {person_clean}")
                        self.logger.debug(f"  檢查 {person_clean} 是否在 name_to_row_idx 中: {person_clean in name_to_row_idx}")
                        self.logger.debug(f"  檢查 @{person_clean} 是否在 name_to_row_idx 中: {f'@{person_clean}' in name_to_row_idx}")
                        
                        # 找到此人的行
                        row_found = False
                        if person_clean in name_to_row_idx:
                            row_idx = name_to_row_idx[person_clean]
                            row_found = True
                            self.logger.debug(f"  >>> 找到人員 {person_clean} 在第 {row_idx} 行")
                            
                            # 填入配對者
                            for i, partner in enumerate(partners):
                                if i < len(new_columns):  # 確保不會超出新增的列數
                                    col_idx = name_col_idx + 1 + i
                                    self.logger.debug(f"  >>> 寫入Excel配對者欄位: {partner} -> 第{row_idx}行第{col_idx}列")
                                    people_sheet.cell(row=row_idx, column=col_idx).value = partner
                                else:
                                    self.logger.warning(f"  >>> 警告: 配對者 {partner} 超出可用欄位數量")
                        elif f"@{person_clean}" in name_to_row_idx:
                            row_idx = name_to_row_idx[f"@{person_clean}"]
                            row_found = True
                            self.logger.debug(f"  >>> 找到人員 @{person_clean} 在第 {row_idx} 行")
                            
                            # 填入配對者
                            for i, partner in enumerate(partners):
                                if i < len(new_columns):  # 確保不會超出新增的列數
                                    col_idx = name_col_idx + 1 + i
                                    self.logger.debug(f"  >>> 寫入Excel配對者欄位: {partner} -> 第{row_idx}行第{col_idx}列")
                                    people_sheet.cell(row=row_idx, column=col_idx).value = partner
                                else:
                                    self.logger.warning(f"  >>> 警告: 配對者 {partner} 超出可用欄位數量")
                        
                        if not row_found:
                            self.logger.warning(f"  >>> 警告: 在name_to_row_idx中找不到人員 {person_clean} 或 @{person_clean}")
                            self.logger.debug(f"  >>> 此人員將在第二階段作為新人員處理")
                    
                    # 添加新人員（不在現有名單中的人）
                    # 只收集本次真正參與配對的人員（從matches參數中獲取，而不是從配對者中獲取）
                    self.logger.info(f"=== 開始收集本次參與配對的人員 ===")
                    self.logger.debug(f"原始matches參數: {matches}")
                    
                    # 從matches中收集所有參與配對的人員
                    participating_people = set()
                    for match in matches:
                        for person in match:
                            if isinstance(person, str):
                                person_clean = person[1:].strip() if person.startswith('@') else person.strip()
                                self.logger.debug(f"從matches添加參與配對人員: {person} -> 清理後: {person_clean}")
                                participating_people.add(person_clean)
                    
                    self.logger.info(f"本次參與配對的人員: {participating_people}")
                    
                    # 只處理真正參與配對的人員，不處理配對者中可能出現的歷史人員
                    all_people = participating_people

                    # 然後檢查每個人是否已在名單中，如果不在則添加
                    self.logger.info("=== 開始檢查並添加新人員到Excel ===")
                    self.logger.debug(f"需要處理的人員: {all_people}")
                    self.logger.debug(f"現有name_to_row_idx: {name_to_row_idx}")
                    
                    for person_clean in all_people:
                        person_with_at = f"@{person_clean}"
                        
                        # 檢查此人是否在現有名單中
                        is_new_person = person_clean not in name_to_row_idx and person_with_at not in name_to_row_idx
                        self.logger.debug(f"\n處理人員: {person_clean}")
                        self.logger.debug(f"  - person_clean在name_to_row_idx中: {person_clean in name_to_row_idx}")
                        self.logger.debug(f"  - person_with_at在name_to_row_idx中: {person_with_at in name_to_row_idx}")
                        self.logger.debug(f"  - 是否為新人員: {is_new_person}")
                        
                        if is_new_person:
                            self.logger.info(f"  >>> 開始添加新人員: {person_clean}")
                            
                            # 找到實際的最後一行（有數據的）
                            actual_last_row = 1  # 從標題行開始
                            for row in range(1, people_sheet.max_row + 1):
                                if people_sheet.cell(row=row, column=name_col_idx).value:
                                    actual_last_row = row
                            
                            # 新增此人到實際的最後一行之後
                            row_idx = actual_last_row + 1
                            self.logger.debug(f"  >>> 寫入Excel姓名欄位: {person_with_at} -> 第{row_idx}行第{name_col_idx}列")
                            people_sheet.cell(row=row_idx, column=name_col_idx).value = person_with_at
                            
                            # 添加配對者（如果此人在match_dict中有配對者）
                            self.logger.debug(f"  >>> 檢查配對者: person_clean={person_clean}, person_with_at={person_with_at}")
                            self.logger.debug(f"  >>> match_dict中的鍵: {list(match_dict.keys())}")
                            
                            partners_found = False
                            if person_clean in match_dict:
                                partners = match_dict[person_clean]
                                self.logger.debug(f"  >>> 找到配對者(使用person_clean): {partners}")
                                partners_found = True
                                for i, partner in enumerate(partners):
                                    if i < len(new_columns):
                                        col_idx = name_col_idx + 1 + i
                                        self.logger.debug(f"  >>> 寫入Excel配對者欄位: {partner} -> 第{row_idx}行第{col_idx}列")
                                        people_sheet.cell(row=row_idx, column=col_idx).value = partner
                            elif person_with_at in match_dict:
                                partners = match_dict[person_with_at]
                                self.logger.debug(f"  >>> 找到配對者(使用person_with_at): {partners}")
                                partners_found = True
                                for i, partner in enumerate(partners):
                                    if i < len(new_columns):
                                        col_idx = name_col_idx + 1 + i
                                        self.logger.debug(f"  >>> 寫入Excel配對者欄位: {partner} -> 第{row_idx}行第{col_idx}列")
                                        people_sheet.cell(row=row_idx, column=col_idx).value = partner
                            
                            if not partners_found:
                                self.logger.warning(f"  >>> 警告: 在match_dict中找不到 {person_clean} 或 {person_with_at} 的配對者")
                            
                            # 更新映射字典
                            name_to_row_idx[person_clean] = row_idx
                            name_to_row_idx[person_with_at] = row_idx
                            self.logger.debug(f"  >>> 更新映射字典: {person_clean} 和 {person_with_at} -> 第{row_idx}行")
                        else:
                            self.logger.debug(f"  >>> 跳過已存在的人員: {person_clean}")
                else:
                    # 如果工作表不存在，則創建新的
                    people_sheet = workbook.create_sheet('人員名單')
                    
                    # 初始化欄位名稱
                    people_sheet.cell(row=1, column=1).value = '姓名'
                    for i, col_name in enumerate(new_columns):
                        people_sheet.cell(row=1, column=2 + i).value = col_name
                    
                    # 填入所有人員和配對結果
                    row_idx = 2
                    for person in people_list:
                        people_sheet.cell(row=row_idx, column=1).value = f"@{person}"
                        
                        if person in match_dict:
                            for i, partner in enumerate(match_dict[person]):
                                if i < len(new_columns):
                                    people_sheet.cell(row=row_idx, column=2 + i).value = partner
                        
                        row_idx += 1
                
                # 保存參與配對人員工作表
                if '參與配對人員' in workbook.sheetnames:
                    participants_sheet = workbook['參與配對人員']
                    # 參與配對人員工作表保持不變
                else:
                    # 如果不存在，則創建新的
                    participants_sheet = workbook.create_sheet('參與配對人員')
                    participants_sheet.cell(row=1, column=1).value = '姓名'
                
                # 保存工作簿
                workbook.save(self.excel_path)
                
                # 標記重複配對
                if repeated_pairs:
                    # 再次打開檔案來設定樣式
                    workbook = openpyxl.load_workbook(self.excel_path)
                    people_sheet = workbook['人員名單']
                    
                    # 找到姓名列和新配對欄位的索引
                    name_col_idx = None
                    new_col_indices = []
                    
                    for col_idx, cell in enumerate(people_sheet[1], 1):
                        if cell.value == '姓名':
                            name_col_idx = col_idx
                        elif cell.value in new_columns:
                            new_col_indices.append(col_idx)
                    
                    if name_col_idx is None:
                        name_col_idx = 1
                    
                    # 設定重複配對的樣式
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    red_font = Font(color="FF0000", bold=True)
                    
                    self.logger.info(f"將檢查以下新配對欄位中的重複配對: {new_columns}")
                    self.logger.debug(f"新配對欄位索引: {new_col_indices}")
                    
                    # 遍歷每一行
                    for row_idx in range(2, people_sheet.max_row + 1):
                        person = people_sheet.cell(row=row_idx, column=name_col_idx).value
                        if not person:
                            continue
                        
                        # 移除 @ 前綴進行比較
                        if isinstance(person, str) and person.startswith('@'):
                            person_clean = person[1:]
                        else:
                            person_clean = person
                        
                        # 遍歷新配對欄位
                        for col_idx in new_col_indices:
                            partner = people_sheet.cell(row=row_idx, column=col_idx).value
                            if not partner:
                                continue
                            
                            # 移除 @ 前綴進行比較
                            partner_norm = partner
                            if isinstance(partner, str):
                                if partner.startswith('@'):
                                    partner_norm = partner[1:].strip()
                                else:
                                    partner_norm = partner.strip()
                            else:
                                partner_norm = str(partner).strip()
                            
                            # 獲取不帶 @ 且去除空格的人名
                            person_norm = person_clean.strip() if isinstance(person_clean, str) else str(person_clean).strip()
                            partner_norm = partner_norm.strip()
                            
                            # 記錄調試信息
                            self.logger.debug(f"檢查是否重複配對: {person_norm} - {partner_norm}")
                            self.logger.debug(f"重複配對列表: {repeated_pairs}")
                            
                            # 檢查是否為重複配對
                            for pair in repeated_pairs:
                                pair_set = set(pair)  # 轉換為集合便於比較
                                if person_norm in pair_set and partner_norm in pair_set:
                                    # 這是重複配對，設定黃底紅字
                                    cell = people_sheet.cell(row=row_idx, column=col_idx)
                                    cell.fill = yellow_fill
                                    cell.font = red_font
                    
                    # 保存工作簿
                    workbook.save(self.excel_path)

            except Exception as e:
                self.logger.error(f"更新人員名單時出錯: {str(e)}")
                # 如果讀取或處理現有資料失敗，就創建新的檔案（原有的邏輯）
                
                # 收集所有參與配對的人員（包括配對者和被配對者）
                all_people = set()
                for person, partners in match_dict.items():
                    if isinstance(person, str):
                        person_clean = person[1:].strip() if person.startswith('@') else person.strip()
                        all_people.add(person_clean)
                    
                    # 添加所有配對者
                    for partner in partners:
                        if isinstance(partner, str):
                            partner_clean = partner[1:].strip() if partner.startswith('@') else partner.strip()
                            all_people.add(partner_clean)
                
                # 使用 all_people 替代原來的方法
                people_list = sorted(list(all_people))
                
                # 創建 DataFrame
                # 人名前加上 @
                people_data = {'姓名': [f"@{person}" for person in people_list]}
                
                # 添加配對者欄位
                for i in range(max_partners):
                    if max_partners > 1:
                        column_name = f"配對者 {today} {i+1}"
                    else:
                        column_name = f"配對者 {today}"
                    
                    # 將配對者添加到對應欄位
                    partners_column = []
                    for person in people_list:
                        person_with_at = f"@{person}"
                        found_partners = []
                        
                        # 檢查各種版本的名稱
                        if person in match_dict and i < len(match_dict[person]):
                            found_partners = match_dict[person][i]
                        elif person_with_at in match_dict and i < len(match_dict[person_with_at]):
                            found_partners = match_dict[person_with_at][i]
                        
                        partners_column.append(found_partners if found_partners else '')
                    
                    people_data[column_name] = partners_column
                
                people_df = pd.DataFrame(people_data)
                
                # 在寫入 Excel 前，保留原始參與配對人員
                try:
                    # 先嘗試讀取現有的參與配對人員
                    existing_participants_df = pd.read_excel(self.excel_path, sheet_name='參與配對人員')
                except:
                    # 如果讀取失敗，則使用空的 DataFrame
                    existing_participants_df = pd.DataFrame(columns=['姓名'])
                
                # 在寫入 Excel 時，使用現有的參與配對人員 DataFrame
                with pd.ExcelWriter(self.excel_path) as writer:
                    people_df.to_excel(writer, sheet_name='人員名單', index=False)
                    existing_participants_df.to_excel(writer, sheet_name='參與配對人員', index=False)
            
        except FileNotFoundError:
            # 如果檔案不存在，創建新的檔案
            import pandas as pd
            
            # 創建配對結果字典，方便查詢每個人的配對者
            match_dict = {}
            for match in matches:
                if len(match) == 2:
                    # 確保每個人的鍵存在
                    if match[0] not in match_dict:
                        match_dict[match[0]] = []
                    if match[1] not in match_dict:
                        match_dict[match[1]] = []
                    
                    # 添加配對關係（確保只有一個 @ 前綴）
                    partner1 = match[1]
                    if not isinstance(partner1, str):
                        partner1_with_at = f"@{str(partner1)}"
                    elif partner1.startswith('@'):
                        partner1_with_at = partner1  # 已有 @ 前綴，保持不變
                    else:
                        partner1_with_at = f"@{partner1}"
                    
                    partner2 = match[0]
                    if not isinstance(partner2, str):
                        partner2_with_at = f"@{str(partner2)}"
                    elif partner2.startswith('@'):
                        partner2_with_at = partner2  # 已有 @ 前綴，保持不變
                    else:
                        partner2_with_at = f"@{partner2}"
                    
                    match_dict[match[0]].append(partner1_with_at)
                    match_dict[match[1]].append(partner2_with_at)
                elif len(match) == 3:
                    # 確保每個人的鍵存在
                    if match[0] not in match_dict:
                        match_dict[match[0]] = []
                    if match[1] not in match_dict:
                        match_dict[match[1]] = []
                    if match[2] not in match_dict:
                        match_dict[match[2]] = []
                    
                    # 添加配對關係（確保只有一個 @ 前綴）
                    for i in range(3):
                        for j in range(3):
                            if i != j:  # 避免自己配對自己
                                person = match[i]
                                partner = match[j]
                                
                                # 確保 partner 只有一個 @ 前綴
                                if not isinstance(partner, str):
                                    partner_with_at = f"@{str(partner)}"
                                elif partner.startswith('@'):
                                    partner_with_at = partner  # 已有 @ 前綴，保持不變
                                else:
                                    partner_with_at = f"@{partner}"
                                
                                match_dict[person].append(partner_with_at)
            
            # 創建人員名單 DataFrame
            # 收集所有參與配對的人員（包括配對者和被配對者）
            all_people = set()
            for person, partners in match_dict.items():
                if isinstance(person, str):
                    person_clean = person[1:].strip() if person.startswith('@') else person.strip()
                    all_people.add(person_clean)
                
                # 添加所有配對者
                for partner in partners:
                    if isinstance(partner, str):
                        partner_clean = partner[1:].strip() if partner.startswith('@') else partner.strip()
                        all_people.add(partner_clean)
            
            # 使用 all_people 替代原來的方法
            people_list = sorted(list(all_people))
            
            import datetime
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            
            # 找出最大配對者數量
            max_partners = max([len(partners) for partners in match_dict.values()], default=1)
            
            # 創建 DataFrame
            # 人名前加上 @
            people_data = {'姓名': [f"@{person}" for person in people_list]}
            self.logger.debug(f"FileNotFoundError處理: 寫入人員名單A欄位的人員: {[f'@{person}' for person in people_list]}")

            
            # 添加配對者欄位
            for i in range(max_partners):
                if max_partners > 1:
                    column_name = f"配對者 {today} {i+1}"
                else:
                    column_name = f"配對者 {today}"
                
                # 將配對者添加到對應欄位
                partners_column = []
                for person in people_list:
                    person_with_at = f"@{person}"
                    found_partners = []
                    
                    # 檢查各種版本的名稱
                    if person in match_dict and len(match_dict[person]) > i:
                        found_partners = match_dict[person][i]
                    elif person_with_at in match_dict and len(match_dict[person_with_at]) > i:
                        found_partners = match_dict[person_with_at][i]
                    
                    partners_column.append(found_partners if found_partners else '')
                
                people_data[column_name] = partners_column
            
            people_df = pd.DataFrame(people_data)
            
            # 創建參與配對人員 DataFrame
            participants_df = pd.DataFrame(columns=['姓名'])
            
            with pd.ExcelWriter(self.excel_path) as writer:
                people_df.to_excel(writer, sheet_name='人員名單', index=False)
                participants_df.to_excel(writer, sheet_name='參與配對人員', index=False)
            
            # 註意：如果是新建檔案，需要再次打開來設定樣式
            if repeated_pairs:
                # 再次打開檔案來設定樣式
                workbook = openpyxl.load_workbook(self.excel_path)
                people_sheet = workbook['人員名單']
                
                # 找到姓名列的索引
                name_col_idx = None
                for col_idx, cell in enumerate(people_sheet[1], 1):
                    if cell.value == '姓名':
                        name_col_idx = col_idx
                        break
                
                if name_col_idx is None:
                    name_col_idx = 1
                
                # 設定重複配對的樣式
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                red_font = Font(color="FF0000", bold=True)
                
                # 只遍歷新配對欄位
                new_col_indices = []
                for col_idx, cell in enumerate(people_sheet[1], 1):
                    if cell.value in new_columns:
                        new_col_indices.append(col_idx)
                
                self.logger.info(f"將檢查以下新配對欄位中的重複配對: {new_columns}")
                self.logger.debug(f"新配對欄位索引: {new_col_indices}")
                
                # 遍歷每一行
                for row_idx in range(2, people_sheet.max_row + 1):
                    person = people_sheet.cell(row=row_idx, column=name_col_idx).value
                    if not person:
                        continue
                    
                    # 移除 @ 前綴進行比較
                    if isinstance(person, str) and person.startswith('@'):
                        person_clean = person[1:]
                    else:
                        person_clean = person
                    
                    # 遍歷新配對欄位
                    for col_idx in new_col_indices:
                        partner = people_sheet.cell(row=row_idx, column=col_idx).value
                        if not partner:
                            continue
                        
                        # 移除 @ 前綴進行比較
                        partner_norm = partner
                        if isinstance(partner, str):
                            if partner.startswith('@'):
                                partner_norm = partner[1:].strip()
                            else:
                                partner_norm = partner.strip()
                        else:
                            partner_norm = str(partner).strip()
                        
                        # 獲取不帶 @ 且去除空格的人名
                        person_norm = person_clean.strip() if isinstance(person_clean, str) else str(person_clean).strip()
                        partner_norm = partner_norm.strip()
                        
                        # 記錄調試信息
                        self.logger.debug(f"檢查是否重複配對: {person_norm} - {partner_norm}")
                        self.logger.debug(f"重複配對列表: {repeated_pairs}")
                        
                        # 檢查是否為重複配對
                        for pair in repeated_pairs:
                            pair_set = set(pair)  # 轉換為集合便於比較
                            if person_norm in pair_set and partner_norm in pair_set:
                                # 這是重複配對，設定黃底紅字
                                cell = people_sheet.cell(row=row_idx, column=col_idx)
                                cell.fill = yellow_fill
                                cell.font = red_font
                
                # 保存工作簿
                workbook.save(self.excel_path)

    def is_valid_pair(self, pair: Tuple[str, ...], history: Set[Tuple[str, ...]]) -> bool:
        """
        檢查配對是否有效
        - 檢查所有可能的2人和3人子組合是否出現在歷史記錄中
        """
        # 標準化 pair 中的所有名稱（去除 @ 前綴）
        normalized_pair = []
        for name in pair:
            norm_name = name
            if isinstance(name, str):
                if name.startswith('@'):
                    norm_name = name[1:].strip()
                else:
                    norm_name = name.strip()
            else:
                norm_name = str(name).strip()
            normalized_pair.append(norm_name)
        
        self.logger.debug(f"檢查配對是否有效: {normalized_pair}")
        
        # 將 pair 中的所有可能 2 人組合檢查是否在歷史記錄中
        for combo in combinations(normalized_pair, 2):
            sorted_combo = tuple(sorted(combo))
            self.logger.debug(f"檢查子配對: {sorted_combo}")
            
            # 直接檢查是否在歷史記錄中
            if sorted_combo in history:
                self.logger.info(f"發現重複配對 (直接匹配): {sorted_combo}")
                return False
            
            # 更嚴格的檢查：遍歷歷史記錄中的每一對
            for hist_pair in history:
                if len(hist_pair) != 2:
                    continue
                    
                # 標準化歷史配對中的名稱
                hist_names = []
                for p in hist_pair:
                    p_norm = p.strip() if isinstance(p, str) else str(p).strip()
                    hist_names.append(p_norm)
                
                # 轉換為集合進行比較
                hist_set = set(hist_names)
                combo_set = set(sorted_combo)
                
                if hist_set == combo_set:
                    self.logger.info(f"發現重複配對 (集合比較): {sorted_combo} vs {hist_names}")
                    return False
            
        # 如果是 3 人組合，還需要檢查完整的組合
        if len(normalized_pair) == 3:
            sorted_pair = tuple(sorted(normalized_pair))
            if sorted_pair in history:
                self.logger.info(f"發現重複的三人組: {sorted_pair}")
                return False
            
        return True

    def match_people(self) -> Tuple[List[Tuple[str, ...]], List[Tuple[str, ...]]]:
        """
        配對人員並返回配對結果和重複配對列表
        返回: (matches, repeated_pairs)
        """
        # 從「參與配對人員」分頁獲取本次參與配對的人員
        try:
            participants_df = pd.read_excel(self.excel_path, sheet_name='參與配對人員')
            # 直接獲取人名，不需要移除 @ 前綴
            people = [name for name in participants_df['姓名'].dropna().tolist() if isinstance(name, str)]
            
            # 移除可能的重複人員
            people = list(dict.fromkeys(people))
            
            # 檢查人員名單中是否有重複
            name_set = set()
            for name in people:
                name_normalized = name[1:] if name.startswith('@') else name
                name_normalized = name_normalized.strip()
                if name_normalized in name_set:
                    self.logger.warning(f"警告：人員名單中有重複: {name_normalized}")
                name_set.add(name_normalized)
            
        except Exception as e:
            # 如果讀取失敗，顯示錯誤訊息
            raise Exception(f"無法讀取參與配對人員: {str(e)}")
        
        if not people:
            raise Exception("參與配對人員名單為空")
        
        # 獲取歷史配對記錄（已經處理了 @ 前綴）
        history = self.get_matching_history()
        
        # 找出一個配對方案中的重複配對
        def find_repeated_pairs(matches: List[Tuple[str, ...]]) -> List[Tuple[str, ...]]:
            repeated = []
            
            # 記錄歷史配對記錄
            self.logger.debug(f"檢查重複配對方案，歷史記錄數量: {len(history)}")
            self.logger.debug(f"歷史配對記錄內容: {history}")
            
            # 如果沒有歷史記錄，直接返回空列表
            if not history:
                self.logger.info("無歷史配對記錄，跳過重複配對檢測")
                return repeated
            
            # 先將歷史配對標準化，確保更準確的比較
            standardized_history = set()
            for hist_pair in history:
                std_pair = []
                for name in hist_pair:
                    if not isinstance(name, str):
                        std_name = str(name).strip()
                    elif name.startswith('@'):
                        std_name = name[1:].strip()
                    else:
                        std_name = name.strip()
                    std_pair.append(std_name)
                standardized_history.add(tuple(sorted(std_pair)))
            
            self.logger.debug(f"標準化後的歷史配對: {standardized_history}")
                
            for match in matches:
                # 檢查配對中是否有重複的人
                if len(set(match)) != len(match):
                    self.logger.warning(f"警告：配對中有重複的人: {match}")
                    continue
                
                # 標準化配對中的人名（去除 @ 前綴）
                normalized_match = []
                for name in match:
                    name_norm = name
                    if isinstance(name, str):
                        if name.startswith('@'):
                            name_norm = name[1:].strip()
                        else:
                            name_norm = name.strip()
                    else:
                        name_norm = str(name).strip()
                    normalized_match.append(name_norm)
                
                self.logger.debug(f"檢查配對方案: {normalized_match}")
                
                # 檢查所有可能的2人組合
                for i in range(len(normalized_match)):
                    for j in range(i+1, len(normalized_match)):
                        person1 = normalized_match[i]
                        person2 = normalized_match[j]
                        pair_to_check = tuple(sorted([person1, person2]))
                        
                        self.logger.debug(f"檢查配對組合: {pair_to_check}")
                        
                        # 方法1: 直接檢查標準化後的歷史記錄
                        if pair_to_check in standardized_history:
                            self.logger.warning(f"!!!找到重複配對 (標準化比較): {pair_to_check}!!!")
                            if pair_to_check not in repeated:
                                repeated.append(pair_to_check)
                            continue
                        
                        # 方法2: 直接檢查原始歷史記錄
                        if pair_to_check in history:
                            self.logger.warning(f"!!!找到重複配對 (原始比較): {pair_to_check}!!!")
                            if pair_to_check not in repeated:
                                repeated.append(pair_to_check)
                            continue
                            
                        # 方法3: 更全面的檢查，遍歷歷史記錄並進行集合比較
                        for hist_pair in history:
                            if len(hist_pair) != 2:
                                continue
                                
                            # 標準化歷史配對
                            hist_names = []
                            for p in hist_pair:
                                p_norm = p.strip() if isinstance(p, str) else str(p).strip()
                                hist_names.append(p_norm)
                            
                            # 轉換為集合進行比較
                            hist_set = set(hist_names)
                            pair_set = set(pair_to_check)
                            
                            # 檢查兩個集合是否相同
                            if pair_set == hist_set:
                                self.logger.warning(f"!!!找到重複配對 (集合比較): {pair_to_check} vs {hist_names}!!!")
                                if pair_to_check not in repeated:
                                    repeated.append(pair_to_check)
                                break
            
            self.logger.info(f"最終重複配對列表: {repeated}")
            return repeated
        
        # 計算一個配對方案中的重複配對數量
        def count_repeated_pairs(matches: List[Tuple[str, ...]]) -> int:
            return len(find_repeated_pairs(matches))
        
        # 使用回溯法找出所有可能的配對方案
        def find_all_matchings(remaining: List[str], current_matches: List[Tuple[str, ...]]) -> List[List[Tuple[str, ...]]]:
            if not remaining:  # 基本情況：沒有剩餘的人要配對
                return [current_matches]
                
            all_matchings = []
            
            if len(remaining) == 2:  # 只剩兩個人
                pair = tuple(sorted(remaining))
                all_matchings.extend(find_all_matchings([], current_matches + [pair]))
                
            elif len(remaining) == 3:  # 只剩三個人
                trio = tuple(sorted(remaining))
                all_matchings.extend(find_all_matchings([], current_matches + [trio]))
                
            else:  # 至少有4個人，可以選擇2人一組
                # 固定第一個人，嘗試與其他每個人配對
                first_person = remaining[0]
                new_remaining = remaining[1:]
                
                for i in range(len(new_remaining)):
                    second_person = new_remaining[i]
                    pair = tuple(sorted([first_person, second_person]))
                    
                    # 準備下一輪遞迴的剩餘人員名單
                    next_remaining = new_remaining.copy()
                    next_remaining.pop(i)
                    
                    # 遞迴找尋剩餘人員的所有可能配對
                    pair_matchings = find_all_matchings(next_remaining, current_matches + [pair])
                    all_matchings.extend(pair_matchings)
            
            return all_matchings
        
        # 主要配對邏輯
        # 先嘗試找出沒有重複配對的方案（提早終止條件）
        def try_no_repeats(remaining: List[str], current_matches: List[Tuple[str, ...]]) -> Tuple[bool, List[Tuple[str, ...]]]:
            if not remaining:
                return True, current_matches
            
            if len(remaining) == 2:
                pair = tuple(sorted(remaining))
                if self.is_valid_pair(pair, history):
                    return True, current_matches + [pair]
                return False, current_matches
                
            if len(remaining) == 3:
                trio = tuple(sorted(remaining))
                if self.is_valid_pair(trio, history):
                    return True, current_matches + [trio]
                return False, current_matches
            
            # 試著先匹配沒有歷史記錄的配對
            first_person = remaining[0]
            new_remaining = remaining[1:]
            
            # 隨機打亂以增加找到解的可能性
            random.shuffle(new_remaining)
            
            for i in range(len(new_remaining)):
                second_person = new_remaining[i]
                pair = tuple(sorted([first_person, second_person]))
                
                if self.is_valid_pair(pair, history):
                    next_remaining = new_remaining.copy()
                    next_remaining.pop(i)
                    
                    success, matches = try_no_repeats(next_remaining, current_matches + [pair])
                    if success:
                        return True, matches
            
            return False, current_matches
        
        # 首先嘗試找到一個無重複的方案（這比窮舉要快得多）
        for _ in range(100):  # 多試幾次隨機順序
            random.shuffle(people)
            success, matches = try_no_repeats(people, [])
            if success:
                return matches, []  # 無重複配對
        
        # 如果人數超過特定閾值，直接使用次優解方案
        if len(people) > 10:  # 根據實際需求調整閾值
            self.logger.info("參與人數過多，使用啟發式方法尋找次優解...")
            
            best_solution = None
            best_score = float('inf')
            fallback_attempts = 1000  # 增加嘗試次數以找到更好的解
            
            for _ in range(fallback_attempts):
                all_people = people.copy()
                random.shuffle(all_people)
                matches = []
                
                while len(all_people) >= 2:
                    if len(all_people) == 2:
                        matches.append(tuple(sorted(all_people)))
                        all_people = []
                    elif len(all_people) == 3:
                        matches.append(tuple(sorted(all_people)))
                        all_people = []
                    else:
                        matches.append(tuple(sorted(all_people[:2])))
                        all_people = all_people[2:]
                
                score = count_repeated_pairs(matches)
                
                if score < best_score:
                    best_score = score
                    best_solution = matches
                    
                    if score == 0:  # 找到無重複解，立即返回
                        repeated_pairs = find_repeated_pairs(best_solution)
                        return best_solution, repeated_pairs
            
            self.logger.info(f"已找到最佳次優解決方案，重複配對數: {best_score}")
            repeated_pairs = find_repeated_pairs(best_solution)
            return best_solution, repeated_pairs
        
        # 對於人數較少的情況，使用窮舉法尋找所有可能的配對方案
        self.logger.info("開始窮舉所有可能的配對方案...")
        all_possible_matchings = find_all_matchings(people, [])
        
        self.logger.info(f"共找到 {len(all_possible_matchings)} 種可能的配對方案")
        
        # 找出重複配對最少的方案
        best_matching = None
        min_repeats = float('inf')
        
        for matching in all_possible_matchings:
            repeats = count_repeated_pairs(matching)
            if repeats < min_repeats:
                min_repeats = repeats
                best_matching = matching
                
                # 如果找到完全無重複的方案，立即返回
                if repeats == 0:
                    self.logger.info("找到了無重複的配對方案！")
                    return best_matching, []  # 無重複配對
        
        if best_matching:
            self.logger.info(f"已找到最佳配對方案，重複配對數: {min_repeats}")
            repeated_pairs = find_repeated_pairs(best_matching)
            return best_matching, repeated_pairs
        else:
            raise Exception("無法完成配對，請管理員手動調整")

def main():
    # 使用特殊方式啟動 TK 應用程式，避免 macOS 顯示終端機窗口
    app = MatchingGUI()
    
    # 在 macOS 上設置應用程式圖標並提高進程優先級
    if sys.platform == 'darwin':
        try:
            # 隱藏終端機窗口
            os.system('''/usr/bin/osascript -e 'tell app "Finder" to set frontmost of process "python" to false' ''')
            
            # 提高進程優先級
            import subprocess
            subprocess.call(['/usr/bin/defaults', 'write', 
                            'com.apple.dock', 'workspaces-auto-swoosh', 
                            '-bool', 'NO'])
            subprocess.call(['/usr/bin/killall', 'Dock'])
        except:
            pass
    
    app.run()

# 使用專門的 macOS 應用程式入口點
if __name__ == "__main__":
    # 檢測是否在 macOS 上運行的打包應用
    if sys.platform == 'darwin' and getattr(sys, 'frozen', False):
        # 改變工作目錄到應用程式包內
        os.chdir(os.path.dirname(os.path.abspath(sys.executable)))
        
        # 隱藏 dock 圖標
        try:
            from AppKit import NSBundle
            bundle = NSBundle.mainBundle()
            info = bundle.localizedInfoDictionary() or bundle.infoDictionary()
            if info and info['CFBundleName'] == 'Python':
                info['LSUIElement'] = '1'  # 設置為後台應用
        except:
            pass
    
    main()
